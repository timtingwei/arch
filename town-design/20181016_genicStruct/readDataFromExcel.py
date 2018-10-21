#-*- coding: utf-8 -*-

"""
from openpyxl import load_workbook

def getColumnValueList(booksheet, column, start_row):
    # 读取一个booksheet的某一列的值, 可以标记头行
    lst = []
    row = start_row
    while (1):
        value = booksheet.cell(row, column).value
        if (value == None):
            break
        lst.append(value)
        row = row + 1
    return lst



if __name__ == '__main__':
    wb = load_workbook('/Users/htwt/Desktop/20181019_totalGenics.xlsx')
    sheets = wb.sheetnames
    # 表头
    sheet_head = sheets[0]
    booksheet = wb[sheet_head]
    tot_arch_type = getColumnValueList(booksheet, 5, 2)   # 所有建筑类型
    tot_job_type = getColumnValueList(booksheet, 6, 2)    # 所有职业类型
    tot_trans_type = getColumnValueList(booksheet, 7, 2)  # 所有出行方式类型
    tot_trans_speed  = getColumnValueList(booksheet, 8, 2)  # 所有出行方式速度

    genicDataLst_lst = []   # 存放所有职业的数据list的list
    for sheetName in wb.sheetnames[1:]:
        lst = []
        booksheet = wb[sheetName]
        jobName = getColumnValueList(booksheet, 1, 3)    # 该职业名字
        lst.append(jobName)
        # 每一个分类结构放在一个list下
        for i in range(15)[2:14:4]:  # [2, 6, 10, 14]
            activity = getColumnValueList(booksheet, i, 3)
            domain = getColumnValueList(booksheet, i+1, 3)
            place = getColumnValueList(booksheet, i+2, 3)
            isDup = getColumnValueList(booksheet, i+3, 3)
            node_lst = [activity, domain, place, isDup]
            lst.append(node_lst)
        breakup = getColumnValueList(booksheet, 18, 3)
        trans_type = getColumnValueList(booksheet, 19, 3)
        lst.append(breakup)
        lst.append(trans_type)
        
        genicDataLst_lst.append(lst)
    print(genicDataLst_lst)
    
"""

'Python 2.7'
import sys
reload(sys) # Python2.5 初始化后会删除 sys.setdefaultencoding 这个方法，我们需要重新载入 
sys.setdefaultencoding('utf-8') 
from xlrd import open_workbook
import string

def getColumnValueList(booksheet, column, start_row):
    # 读取一个booksheet的某一列的值, 可以标记头行
    lst = []
    row = start_row
    value = ''
    nrows = booksheet.nrows
    for row in range(start_row, nrows+1):
        value = booksheet.cell(row-1, column-1).value
        if (value == ''):  # 判断空
            break
        if isinstance(value, unicode):
            value = value.encode('utf-8')
        lst.append(value)
    return lst
"""
def main():
    wb = open_workbook('/Users/htwt/Desktop/20181019_totalGenics.xls')
    tot_sheets = wb.sheets()
    booksheet = tot_sheets[0]
    ncols = booksheet.ncols
    print([str(s.encode('utf-8')) for s in booksheet.col_values(3)])
"""


def read(filepath):
    wb = open_workbook(filepath)
    tot_sheets = wb.sheets()
    # 表头
    #sheet_head = sheets[0]
    booksheet = tot_sheets[0]
    tot_job_scale = getColumnValueList(booksheet, 2, 2)   # 所有职业人数分配比例
    tot_block_num = getColumnValueList(booksheet, 4, 2)   # 所有建筑类型对应区块的数量
    tot_arch_type = getColumnValueList(booksheet, 5, 2)   # 所有建筑类型
    tot_job_type = getColumnValueList(booksheet, 6, 2)    # 所有职业类型
    tot_trans_type = getColumnValueList(booksheet, 7, 2)  # 所有出行方式类型
    tot_trans_speed  = getColumnValueList(booksheet, 8, 2)  # 所有出行方式速度

    genicDataLst_lst = []   # 存放所有职业的数据list的list
    for booksheet in tot_sheets[1:]:
        lst = []
        jobName = getColumnValueList(booksheet, 1, 3)    # 该职业名字
        lst.append(jobName)
        # 每一个分类结构放在一个list下
        for i in range(15)[2:15:4]:  # [2, 6, 10, 14]
            activity = getColumnValueList(booksheet, i, 3)
            domain = getColumnValueList(booksheet, i+1, 3)
            place = getColumnValueList(booksheet, i+2, 3)
            isDup = getColumnValueList(booksheet, i+3, 3)
            node_lst = [activity, domain, place, isDup]
            lst.append(node_lst)
        breakup = getColumnValueList(booksheet, 18, 3)
        trans_type = getColumnValueList(booksheet, 19, 3)
        lst.append(breakup)
        lst.append(trans_type)
        
        genicDataLst_lst.append(lst)
    return genicDataLst_lst, tot_job_scale, tot_block_num, tot_arch_type, tot_job_type, tot_trans_type, tot_trans_speed
    #print(str(genicDataLst_lst).decode('string_escape'))
    #print(genicDataLst_lst)[-9][1][0][0]
    #print(genicDataLst_lst[0])
    #print(len(genicDataLst_lst[0]))


if __name__ == '__main__':
    read('/Users/htwt/Desktop/20181019_totalGenics.xls')
